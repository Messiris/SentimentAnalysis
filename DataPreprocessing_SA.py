import pathlib
import csv
import openpyxl
import re
import os
import json
import warnings
import pymorphy2


class DataPreprocessing_SA(object):

    dir_path = pathlib.Path.cwd()

    def __init__(self, data = "doc_comment_summary.xlsx"):
        self.sentimentdata_file_path = pathlib.Path(self.dir_path, "data", data)

    def deleteTrain(self, train = "train.json"):
        train_path = pathlib.Path(self.dir_path, "data", train)
        if os.path.exists(train_path):
            try:
                os.remove(train_path)
                print("Train deleted")
            except PermissionError:
                print("Failed to deletete train - already opened")

    def writeTrain(self, train='train', nstr=0):
        wookbook = openpyxl.load_workbook(self.sentimentdata_file_path)
        worksheet = wookbook.active
        lemmatizer = pymorphy2.MorphAnalyzer()
        if nstr == 0:
            nstr = worksheet.max_row
        for row in worksheet.iter_rows(1, nstr):
            if isinstance(row[1].value, int) and isinstance(row[0].value, str):
                # print(row[0].value)
                # print(row[1].value)
                ctext = row[0].value
                ctext = ctext.lower()
                ctext = re.sub("ё", "е", ctext)
                ctext = re.sub("[^а-я]+", " ", ctext)
                ctext = ctext.split(' ')
                text = ''
                for tx in ctext:
                    txt = lemmatizer.normal_forms(tx)
                    text = text + ' ' + txt[0]
                with open('data/' + train + '.json', 'a') as f:
                    json.dump({'text': text, 'rating': row[1].value}, f, ensure_ascii=False)
                    f.write('\n')
            else:
                print(row, row[1].value, row[0].value)

